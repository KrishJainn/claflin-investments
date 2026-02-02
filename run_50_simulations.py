#!/usr/bin/env python3
"""
AQTIS Multi-Run Backtest with Agent-Driven Learning.

Runs N iterations of 60-day backtests using the PaperTrader engine,
which routes EVERY decision through the full agent stack:
  - Orchestrator lightweight_signal_check (memory-based filtering)
  - Pre-trade workflow for high-conviction signals (Gemini LLM)
  - Risk manager position sizing + adaptive exits
  - Batch post-mortem learning every 10 days
  - Periodic review every 3 days
  - Mid-run strategy review at midpoint

Each run bootstraps from the best prior snapshot, so weights evolve
across runs via agent intelligence (not external WeightEvolver).

Gemini API is wired into EVERY agent through the orchestrator.

Usage:
    python run_50_simulations.py
    python run_50_simulations.py --runs 10
    python run_50_simulations.py --runs 50 --days 30
"""

import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import numpy as np
import pandas as pd

# Load .env for API keys
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

# Project root
sys.path.insert(0, str(Path(__file__).parent))

from aqtis.config.settings import load_config
from aqtis.memory.memory_layer import MemoryLayer
from aqtis.llm.base import MockLLMProvider
from aqtis.orchestrator.orchestrator import MultiAgentOrchestrator
from aqtis.backtest.paper_trader import PaperTrader

logger = logging.getLogger(__name__)


def _get_llm():
    """Initialize the best available LLM provider (Gemini preferred)."""
    api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    if api_key:
        try:
            from aqtis.llm.gemini_provider import GeminiProvider
            provider = GeminiProvider(api_key=api_key)
            if provider.is_available():
                logger.info("Using Gemini LLM (gemini-2.0-flash)")
                return provider, "gemini"
        except Exception as e:
            logger.warning(f"Gemini init failed: {e}")

    logger.warning("No LLM API key found, falling back to MockLLMProvider")
    return MockLLMProvider(
        default_response='{"result": "ok", "should_trade": false, "confidence": 0.5}',
    ), "mock"


def run_multi_simulation(
    num_runs: int = 50,
    days: int = 60,
    llm_budget_per_run: int = 80,
    config_path: str = "aqtis_config.yaml",
):
    """Run multiple simulation iterations with agent-driven cross-run learning."""

    # Setup
    config = load_config(config_path)
    data_dir = Path("aqtis_data")
    data_dir.mkdir(exist_ok=True)

    memory = MemoryLayer(
        db_path=str(data_dir / "aqtis.db"),
        vector_path=str(data_dir / "vectors"),
    )

    # Gemini LLM — wired into EVERY agent via orchestrator
    llm, llm_type = _get_llm()

    symbols = config.market_data.symbols
    initial_capital = config.execution.initial_capital

    print("=" * 65)
    print(f"  AQTIS AGENT-DRIVEN MULTI-RUN BACKTEST")
    print("=" * 65)
    print(f"  Engine:     PaperTrader (all 6 agents active)")
    print(f"  Capital:    ${initial_capital:,.0f} USD")
    print(f"  Symbols:    {len(symbols)} NIFTY stocks")
    print(f"  Days/Run:   {days}")
    print(f"  Total Runs: {num_runs}")
    print(f"  LLM:        {llm_type.upper()} ({llm_budget_per_run} calls/run)")
    print(f"  Learning:   Batch every 10 days, Review every 3 days")
    print("=" * 65)
    print()

    # Results storage
    all_results: List[Dict] = []
    cumulative_pnl = 0.0
    best_sharpe = -999
    best_run = 0
    start_time = time.time()

    for run_num in range(1, num_runs + 1):
        run_start = time.time()

        # Create fresh orchestrator + PaperTrader per run
        # Orchestrator passes LLM to ALL agents
        orchestrator = MultiAgentOrchestrator(
            memory=memory, llm=llm,
            config={"portfolio_value": initial_capital},
        )

        trader = PaperTrader(
            memory=memory,
            orchestrator=orchestrator,
            config=config,
            llm_budget=llm_budget_per_run,
        )

        print(f"\n{'─'*55}")
        print(f"  RUN {run_num}/{num_runs}  |  Cumulative P&L: ${cumulative_pnl:+,.2f}")
        print(f"{'─'*55}")

        # Run agent-driven backtest
        result = trader.run(
            symbols=symbols,
            days=days,
            initial_capital=initial_capital,
            run_number=run_num,
        )

        run_elapsed = time.time() - run_start

        if result.get("error"):
            print(f"  ERROR: {result['error']}")
            all_results.append({
                "run": run_num, "error": result["error"],
                "pnl": 0, "return_pct": 0, "trades": 0,
                "win_rate": 0, "sharpe": 0,
            })
            continue

        cap = result.get("capital", {})
        metrics = result.get("metrics", {})
        llm_usage = result.get("llm_usage", {})
        learning = result.get("learning", {})
        pnl = cap.get("pnl", 0)
        ret_pct = cap.get("return_pct", 0)
        cumulative_pnl += pnl

        sharpe = metrics.get("sharpe_ratio", 0)
        if sharpe > best_sharpe:
            best_sharpe = sharpe
            best_run = run_num

        wr = metrics.get("win_rate", 0)
        trades = metrics.get("total_trades", 0)
        pf = metrics.get("profit_factor", 0)
        mdd = metrics.get("max_drawdown", 0)

        print(f"  P&L: ${pnl:+,.2f} ({ret_pct:+.2f}%)  |  Trades: {trades}")
        print(f"  WR: {wr:.1%}  Sharpe: {sharpe:.2f}  PF: {pf:.2f}  MDD: {mdd:.2%}")
        print(f"  LLM: {llm_usage.get('calls_used', 0)}/{llm_usage.get('budget', 0)}  "
              f"|  Mutations: {learning.get('weight_mutations', 0)}  "
              f"|  Threshold: {learning.get('final_entry_threshold', 0):.3f}")
        print(f"  Time: {run_elapsed:.1f}s")

        # Show agent activity summary
        agent_log = result.get("agent_log", [])
        if agent_log:
            agent_counts = {}
            for entry in agent_log:
                agent = entry.get("agent", "unknown")
                agent_counts[agent] = agent_counts.get(agent, 0) + 1
            agent_str = ", ".join(f"{a}: {c}" for a, c in sorted(agent_counts.items()))
            print(f"  Agents: {agent_str}")

        # Exit reason breakdown
        exit_reasons = metrics.get("exit_reasons", {})
        if exit_reasons:
            er_str = ", ".join(f"{k}: {v}" for k, v in exit_reasons.items())
            print(f"  Exits: {er_str}")

        # Store run results
        regime = metrics.get("regime_breakdown", {})
        regime_str = "; ".join(
            f"{r}: {s['trades']}t/{s['wins']}w" for r, s in regime.items()
        ) if regime else ""

        # Agent actions summary
        agent_actions = []
        for entry in agent_log[-5:]:  # Last 5 actions
            agent_actions.append(
                f"{entry.get('agent', '?')}:{entry.get('action', '?')}"
            )

        all_results.append({
            "run": run_num,
            "pnl": pnl,
            "return_pct": ret_pct,
            "cumulative_pnl": cumulative_pnl,
            "final_capital": cap.get("final", 0),
            "trades": trades,
            "wins": metrics.get("wins", 0),
            "losses": metrics.get("losses", 0),
            "win_rate": wr,
            "sharpe": sharpe,
            "profit_factor": pf,
            "max_drawdown": mdd,
            "avg_win": metrics.get("avg_win", 0),
            "avg_loss": metrics.get("avg_loss", 0),
            "llm_calls": llm_usage.get("calls_used", 0),
            "entry_threshold": learning.get("final_entry_threshold", 0),
            "weight_mutations": learning.get("weight_mutations", 0),
            "regime_breakdown": regime_str,
            "exit_reasons": json.dumps(exit_reasons, default=str),
            "agent_actions": "; ".join(agent_actions),
            "agent_log_count": len(agent_log),
            "elapsed_seconds": run_elapsed,
        })

    # Final summary
    total_elapsed = time.time() - start_time
    print(f"\n{'='*65}")
    print(f"  ALL {num_runs} RUNS COMPLETE")
    print(f"{'='*65}")
    print(f"  Total P&L:      ${cumulative_pnl:+,.2f}")
    print(f"  Best Sharpe:     {best_sharpe:.2f} (Run #{best_run})")
    print(f"  Avg P&L/Run:     ${cumulative_pnl/max(num_runs,1):+,.2f}")
    print(f"  Total Time:      {total_elapsed:.0f}s ({total_elapsed/60:.1f}m)")

    winning_runs = sum(1 for r in all_results if r.get("pnl", 0) > 0)
    losing_runs = sum(1 for r in all_results if r.get("pnl", 0) <= 0)
    print(f"  Winning Runs:    {winning_runs}/{num_runs}")
    print(f"  Losing Runs:     {losing_runs}/{num_runs}")

    pnls = [r.get("pnl", 0) for r in all_results]
    sharpes = [r.get("sharpe", 0) for r in all_results]
    wrs = [r.get("win_rate", 0) for r in all_results]
    avg_sharpe = np.mean(sharpes) if sharpes else 0
    avg_wr = np.mean(wrs) if wrs else 0
    print(f"  Avg Sharpe:      {avg_sharpe:.2f}")
    print(f"  Avg Win Rate:    {avg_wr:.1%}")
    print(f"{'='*65}")

    # Generate Excel
    excel_path = generate_excel(all_results, data_dir)
    print(f"\n  Excel saved: {excel_path}")

    # Save JSON backup
    json_path = data_dir / "multi_run_results.json"
    with open(json_path, "w") as f:
        json.dump(all_results, f, indent=2, default=str)
    print(f"  JSON saved:  {json_path}")

    return all_results


def generate_excel(results: List[Dict], output_dir: Path) -> str:
    """Generate comprehensive Excel summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ─── Sheet 1: Run Summary ───
    ws = wb.active
    ws.title = "Run Summary"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Run #", "P&L ($)", "Return %", "Cumulative P&L ($)",
        "Trades", "Wins", "Losses", "Win Rate",
        "Sharpe", "Profit Factor", "Max Drawdown",
        "Avg Win ($)", "Avg Loss ($)", "LLM Calls",
        "Entry Threshold", "Weight Mutations",
        "Regime Breakdown", "Exit Reasons",
        "Agent Actions", "Agent Events", "Time (s)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        values = [
            r.get("run", ""),
            r.get("pnl", 0),
            r.get("return_pct", 0),
            r.get("cumulative_pnl", 0),
            r.get("trades", 0),
            r.get("wins", 0),
            r.get("losses", 0),
            r.get("win_rate", 0),
            r.get("sharpe", 0),
            r.get("profit_factor", 0),
            r.get("max_drawdown", 0),
            r.get("avg_win", 0),
            r.get("avg_loss", 0),
            r.get("llm_calls", 0),
            r.get("entry_threshold", 0),
            r.get("weight_mutations", 0),
            r.get("regime_breakdown", ""),
            r.get("exit_reasons", ""),
            r.get("agent_actions", ""),
            r.get("agent_log_count", 0),
            r.get("elapsed_seconds", 0),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col == 2:
                if isinstance(val, (int, float)) and val > 0:
                    cell.fill = green_fill
                elif isinstance(val, (int, float)) and val < 0:
                    cell.fill = red_fill
                cell.number_format = "#,##0.00"
            elif col == 3:
                cell.number_format = "0.00%"
            elif col == 4:
                cell.number_format = "#,##0.00"
            elif col == 8:
                cell.number_format = "0.0%"
            elif col in (9, 10):
                cell.number_format = "0.00"
            elif col == 11:
                cell.number_format = "0.00%"
            elif col in (12, 13):
                cell.number_format = "#,##0.00"

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(
            len(headers[col - 1]) + 4, 12
        )
    ws.column_dimensions[get_column_letter(17)].width = 40
    ws.column_dimensions[get_column_letter(18)].width = 40
    ws.column_dimensions[get_column_letter(19)].width = 60

    # ─── Sheet 2: Charts ───
    ws2 = wb.create_sheet("Charts")
    ws2.cell(row=1, column=1, value="Run #").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Cumulative P&L ($)").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Sharpe Ratio").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Win Rate").font = Font(bold=True)
    ws2.cell(row=1, column=5, value="P&L per Run ($)").font = Font(bold=True)

    for i, r in enumerate(results, 2):
        ws2.cell(row=i, column=1, value=r.get("run", 0))
        ws2.cell(row=i, column=2, value=r.get("cumulative_pnl", 0))
        ws2.cell(row=i, column=3, value=r.get("sharpe", 0))
        ws2.cell(row=i, column=4, value=r.get("win_rate", 0))
        ws2.cell(row=i, column=5, value=r.get("pnl", 0))

    chart1 = LineChart()
    chart1.title = "Cumulative P&L Across Runs"
    chart1.x_axis.title = "Run #"
    chart1.y_axis.title = "Cumulative P&L ($)"
    chart1.style = 10
    chart1.width = 25
    chart1.height = 15
    data = Reference(ws2, min_col=2, min_row=1, max_row=len(results) + 1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(results) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.line.width = 25000
    ws2.add_chart(chart1, "G1")

    chart2 = LineChart()
    chart2.title = "Sharpe Ratio Evolution"
    chart2.x_axis.title = "Run #"
    chart2.y_axis.title = "Sharpe Ratio"
    chart2.style = 10
    chart2.width = 25
    chart2.height = 15
    data2 = Reference(ws2, min_col=3, min_row=1, max_row=len(results) + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws2.add_chart(chart2, "G18")

    # ─── Sheet 3: Overall Stats ───
    ws3 = wb.create_sheet("Overall Stats")
    pnls = [r.get("pnl", 0) for r in results]
    sharpes = [r.get("sharpe", 0) for r in results]
    wrs = [r.get("win_rate", 0) for r in results]

    stats = [
        ("Total Runs", len(results)),
        ("Total P&L ($)", sum(pnls)),
        ("Avg P&L per Run ($)", np.mean(pnls) if pnls else 0),
        ("Std P&L ($)", np.std(pnls) if pnls else 0),
        ("Best P&L ($)", max(pnls) if pnls else 0),
        ("Worst P&L ($)", min(pnls) if pnls else 0),
        ("Winning Runs", sum(1 for p in pnls if p > 0)),
        ("Losing Runs", sum(1 for p in pnls if p <= 0)),
        ("Win Rate (Runs)", sum(1 for p in pnls if p > 0) / max(len(pnls), 1)),
        ("", ""),
        ("Avg Sharpe", np.mean(sharpes) if sharpes else 0),
        ("Best Sharpe", max(sharpes) if sharpes else 0),
        ("Avg Win Rate", np.mean(wrs) if wrs else 0),
        ("Best Win Rate", max(wrs) if wrs else 0),
        ("", ""),
        ("First 10 Runs Avg P&L", np.mean(pnls[:10]) if len(pnls) >= 10 else "N/A"),
        ("Last 10 Runs Avg P&L", np.mean(pnls[-10:]) if len(pnls) >= 10 else "N/A"),
        ("First 10 Avg Sharpe", np.mean(sharpes[:10]) if len(sharpes) >= 10 else "N/A"),
        ("Last 10 Avg Sharpe", np.mean(sharpes[-10:]) if len(sharpes) >= 10 else "N/A"),
        ("Improvement (P&L)", (np.mean(pnls[-10:]) - np.mean(pnls[:10])) if len(pnls) >= 20 else "N/A"),
        ("Improvement (Sharpe)", (np.mean(sharpes[-10:]) - np.mean(sharpes[:10])) if len(sharpes) >= 20 else "N/A"),
    ]

    for row_idx, (label, value) in enumerate(stats, 1):
        ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        cell = ws3.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = "#,##0.00"
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15

    # ─── Sheet 4: Agent Activity ───
    ws4 = wb.create_sheet("Agent Activity")
    ws4.cell(row=1, column=1, value="Run #").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="LLM Calls").font = Font(bold=True)
    ws4.cell(row=1, column=3, value="Weight Mutations").font = Font(bold=True)
    ws4.cell(row=1, column=4, value="Agent Events").font = Font(bold=True)
    ws4.cell(row=1, column=5, value="Entry Threshold").font = Font(bold=True)
    ws4.cell(row=1, column=6, value="Agent Actions (last 5)").font = Font(bold=True)

    for i, r in enumerate(results, 2):
        ws4.cell(row=i, column=1, value=r.get("run", 0))
        ws4.cell(row=i, column=2, value=r.get("llm_calls", 0))
        ws4.cell(row=i, column=3, value=r.get("weight_mutations", 0))
        ws4.cell(row=i, column=4, value=r.get("agent_log_count", 0))
        ws4.cell(row=i, column=5, value=r.get("entry_threshold", 0))
        ws4.cell(row=i, column=6, value=r.get("agent_actions", ""))

    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["F"].width = 80

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"AQTIS_Runs_{timestamp}.xlsx"
    filepath = output_dir / filename
    wb.save(str(filepath))

    return str(filepath)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="AQTIS Agent-Driven Multi-Run Backtest")
    parser.add_argument("--runs", type=int, default=50, help="Number of runs (default: 50)")
    parser.add_argument("--days", type=int, default=60, help="Days per run (default: 60)")
    parser.add_argument("--llm-budget", type=int, default=80, help="LLM calls per run (default: 80)")
    parser.add_argument("--config", default="aqtis_config.yaml", help="Config path")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
        datefmt="%H:%M:%S",
    )
    # Suppress noisy loggers
    logging.getLogger("trading_evolution").setLevel(logging.WARNING)
    logging.getLogger("chromadb").setLevel(logging.WARNING)
    logging.getLogger("aqtis.orchestrator").setLevel(logging.WARNING)
    logging.getLogger("aqtis.agents").setLevel(logging.WARNING)

    run_multi_simulation(
        num_runs=args.runs,
        days=args.days,
        llm_budget_per_run=args.llm_budget,
        config_path=args.config,
    )
